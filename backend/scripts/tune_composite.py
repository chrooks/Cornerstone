"""
Quick composite tuning script — compare current vs proposed formula for a set of reference players.

Usage:
    cd backend
    source venv/bin/activate
    python scripts/tune_composite.py
"""

from __future__ import annotations

import os
import sys
from collections import Counter
from dotenv import load_dotenv

load_dotenv()

# Ensure backend is importable
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.supabase_client import get_supabase, run_query
from services.cohesion_engine.composites import compute_raw_composites, tier_value
from services.cohesion_engine.weights import (
    COMPOSITE_COEFFICIENTS,
    COMPOSITE_NAMES,
    THEORETICAL_MAX,
    TIER_VALUES,
    NORMALIZATION_BREAKPOINT_PERCENTILE,
    NORMALIZATION_BREAKPOINT_SCORE,
    MIN_DISTRIBUTION_SIZE,
)
from services.skills import ALL_SKILLS

SEASON = "2025-26"

VALUES = {
    "tier_values": TIER_VALUES,
    "composite_coefficients": COMPOSITE_COEFFICIENTS,
    "theoretical_max": THEORETICAL_MAX,
    "normalization_breakpoint_percentile": NORMALIZATION_BREAKPOINT_PERCENTILE,
    "normalization_breakpoint_score": NORMALIZATION_BREAKPOINT_SCORE,
    "min_distribution_size": MIN_DISTRIBUTION_SIZE,
}


def _extract_skills(profile: dict) -> dict[str, str]:
    return {
        skill: data.get("final_tier", "None") if isinstance(data, dict) else data
        for skill, data in profile.items()
    }


def _with_defaults(skills: dict) -> dict:
    normalized = dict(skills)
    for skill in ALL_SKILLS:
        normalized.setdefault(skill, "None")
    return normalized


def _tv(skills: dict, skill: str) -> float:
    return tier_value(skills, skill, TIER_VALUES)


def load_all_profiles() -> list[dict]:
    """Load all composite profiles + legends with player metadata."""
    client = get_supabase()

    profiles = run_query(
        lambda: client.table("skill_profiles")
        .select("player_id, profile")
        .eq("season", SEASON)
        .eq("source", "composite")
        .execute()
    )
    legend_profiles = run_query(
        lambda: client.table("skill_profiles")
        .select("player_id, profile")
        .eq("source", "manual")
        .eq("is_legend", True)
        .execute()
    )

    all_ids = [r["player_id"] for r in profiles.data + legend_profiles.data if r["player_id"]]
    players_data = run_query(
        lambda: client.table("players").select("id, name, position").in_("id", all_ids).execute()
    )
    pmap = {p["id"]: p for p in players_data.data}

    records = []
    for row in profiles.data:
        if not row["player_id"]:
            continue
        p = pmap.get(row["player_id"], {})
        skills = _with_defaults(_extract_skills(row["profile"]))
        records.append({"name": p.get("name", "?"), "pos": p.get("position", "?"), "skills": skills, "is_legend": False})

    for row in legend_profiles.data:
        if not row["player_id"]:
            continue
        p = pmap.get(row["player_id"], {})
        skills = _with_defaults(_extract_skills(row["profile"]))
        records.append({"name": p.get("name", "?"), "pos": p.get("position", "?"), "skills": skills, "is_legend": True})

    return records


def compute_proposed_shot_creation(skills: dict, raw_composites: dict, coefficients: dict) -> float:
    """Proposed shot_creation formula: pnr_orchestration + passer + od + iso + spacing + paint_touch."""
    c = coefficients
    return (
        c["shot_creation_pnr_orchestration"] * raw_composites["pnr_orchestration"]
        + c.get("shot_creation_passer", 0.0) * _tv(skills, "passer")
        + c["shot_creation_off_dribble"] * _tv(skills, "off_dribble_shooter")
        + _tv(skills, "isolation_scorer")
        + c["shot_creation_spacing"] * raw_composites["spacing"]
        + c["shot_creation_paint_touch"] * raw_composites["paint_touch"]
    )


def pct_normalize(raw: float, distribution: list[float]) -> float:
    """Percentile normalization matching composites.py logic."""
    if raw <= 0 or not distribution:
        return 0.0
    n = len(distribution)
    below = sum(1 for v in distribution if v < raw)
    equal = sum(1 for v in distribution if v == raw)
    percentile = (below + equal / 2) / n
    bp = NORMALIZATION_BREAKPOINT_PERCENTILE
    bs = NORMALIZATION_BREAKPOINT_SCORE
    p_break_index = int(n * bp)
    p_break_value = distribution[min(p_break_index, n - 1)]
    empirical_max = distribution[-1]
    if percentile <= bp:
        return round(percentile / bp * bs, 1)
    elif empirical_max <= p_break_value:
        return 10.0
    else:
        t = max(0.0, min(1.0, (raw - p_break_value) / (empirical_max - p_break_value)))
        return round(min(10.0, bs + t * (10.0 - bs)), 1)


# ---------------------------------------------------------------------------
# Tuning parameters — EDIT THESE
# ---------------------------------------------------------------------------

PROPOSED_COEFFICIENTS = {
    **COMPOSITE_COEFFICIENTS,
    "shot_creation_pnr_orchestration": 0.6,
    "shot_creation_passer": 0.5,
    "shot_creation_off_dribble": 0.7,
    "shot_creation_paint_touch": 0.5,
}

# Reference players to highlight (substring match on name)
REFERENCE_PLAYERS = [
    "Stephen Curry",
    "LeBron James",
    "Nikola Jokić",
    "Luka Dončić",
    "Shai Gilgeous-Alexander",
    "Anthony Edwards",
    "Pascal Siakam",
    "Rudy Gobert",
    "Mikal Bridges",
    "Steven Adams",
    "Kevin Durant",
    "Devin Booker",
    "Tyrese Haliburton",
    "Jayson Tatum",
    "Julius Randle",
    "Cam Thomas",
    "Scottie Barnes",
    "Cade Cunningham",
    "Victor Wembanyama",
    "Trae Young",
    "Jalen Johnson",
    "Amen Thompson",
    "Aaron Gordon",
    "Cooper Flagg",
    "Dylan Harper",
    "Zion Williamson",
]


def export_xlsx(ref_records: list[dict], coefficients: dict) -> None:
    """Export reference player comparison to an xlsx file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Shot Creation Tuning"

    # Header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    # Coefficients section
    ws.append(["Proposed Coefficients"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(["Parameter", "Value"])
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
    coeff_rows = [
        ("shot_creation_pnr_orchestration", coefficients["shot_creation_pnr_orchestration"]),
        ("shot_creation_passer", coefficients.get("shot_creation_passer", "N/A")),
        ("shot_creation_off_dribble", coefficients["shot_creation_off_dribble"]),
        ("shot_creation_spacing", coefficients["shot_creation_spacing"]),
        ("shot_creation_paint_touch", coefficients["shot_creation_paint_touch"]),
        ("isolation_scorer", "1.0 (unweighted)"),
    ]
    for name, val in coeff_rows:
        ws.append([name, val])

    ws.append([])
    ws.append([])

    # Player comparison table
    headers = [
        "Name", "Pos", "Legend",
        "pnr_ball_handler", "passer", "off_dribble_shooter", "isolation_scorer",
        "pnr_orch_raw",
        "Current Raw", "Current Norm", "Proposed Raw", "Proposed Norm", "Δ Norm",
    ]
    header_row = ws.max_row + 1
    ws.append(headers)
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for r in ref_records:
        s = r["skills"]
        delta = round(r["proposed_norm"] - r["current_norm"], 1)
        row = [
            r["name"],
            r["pos"],
            "Yes" if r["is_legend"] else "",
            s.get("pnr_ball_handler", "None"),
            s.get("passer", "None"),
            s.get("off_dribble_shooter", "None"),
            s.get("isolation_scorer", "None"),
            round(r["current_raw"]["pnr_orchestration"], 1),
            round(r["current_shot_creation"], 1),
            round(r["current_norm"], 1),
            round(r["proposed_shot_creation"], 1),
            round(r["proposed_norm"], 1),
            delta,
        ]
        ws.append(row)
        data_row = ws.max_row
        for cell in ws[data_row]:
            cell.border = thin_border
        # Color the delta cell
        delta_cell = ws.cell(row=data_row, column=13)
        if delta > 0:
            delta_cell.fill = green_fill
        elif delta < 0:
            delta_cell.fill = red_fill

    # Column widths
    col_widths = [28, 6, 7, 18, 18, 20, 18, 14, 12, 13, 13, 14, 9]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else ""].width = width

    output_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "..",
        "feature_requests",
        "shot-creation-tuning.xlsx",
    )
    wb.save(output_path)
    print(f"\n📊 Exported to {output_path}")


def main() -> None:
    print("Loading player data...")
    records = load_all_profiles()
    print(f"Loaded {len(records)} players.\n")

    # Compute current and proposed raw values for all players
    current_raw_all = []
    proposed_raw_all = []

    for rec in records:
        raw = compute_raw_composites(rec["skills"], VALUES)
        rec["current_raw"] = raw
        rec["current_shot_creation"] = raw["shot_creation"]

        proposed = compute_proposed_shot_creation(rec["skills"], raw, PROPOSED_COEFFICIENTS)
        rec["proposed_shot_creation"] = proposed

        current_raw_all.append(raw["shot_creation"])
        proposed_raw_all.append(proposed)

    current_dist = sorted(current_raw_all)
    proposed_dist = sorted(proposed_raw_all)

    # Add normalized values
    for rec in records:
        rec["current_norm"] = pct_normalize(rec["current_shot_creation"], current_dist)
        rec["proposed_norm"] = pct_normalize(rec["proposed_shot_creation"], proposed_dist)

    # Print reference players comparison
    print(f"{'Name':<28} {'Pos':>4} {'L':>2} | {'cur_raw':>8} {'cur_norm':>9} | {'new_raw':>8} {'new_norm':>9} | {'Δ_norm':>7}")
    print("-" * 95)

    ref_records = []
    for ref_name in REFERENCE_PLAYERS:
        matches = [r for r in records if ref_name.lower() in r["name"].lower()]
        if matches:
            ref_records.append(matches[0])

    ref_records.sort(key=lambda r: r["proposed_norm"], reverse=True)

    for r in ref_records:
        leg = "L" if r["is_legend"] else ""
        delta = r["proposed_norm"] - r["current_norm"]
        sign = "+" if delta >= 0 else ""
        print(
            f'{r["name"]:<28} {r["pos"]:>4} {leg:>2} | '
            f'{r["current_shot_creation"]:>8.1f} {r["current_norm"]:>9.1f} | '
            f'{r["proposed_shot_creation"]:>8.1f} {r["proposed_norm"]:>9.1f} | '
            f'{sign}{delta:>6.1f}'
        )

    # Distribution comparison
    print("\n=== DISTRIBUTION COMPARISON ===\n")
    print(f"{'Bucket':>6} | {'Current':>8} | {'Proposed':>8}")
    print("-" * 30)
    for b in range(11):
        cur_count = sum(1 for r in records if int(r["current_norm"]) == b)
        new_count = sum(1 for r in records if int(r["proposed_norm"]) == b)
        print(f"  {b:>2}.x  | {cur_count:>8} | {new_count:>8}")

    # Show key Skills for reference players
    print("\n=== REFERENCE PLAYER SKILLS (shot creation inputs) ===\n")
    print(f"{'Name':<28} {'pnr_bh':>8} {'passer':>8} {'od_shoot':>8} {'iso':>8} {'pnr_orch_raw':>13}")
    print("-" * 80)
    for r in ref_records:
        s = r["skills"]
        raw = r["current_raw"]
        print(
            f'{r["name"]:<28} '
            f'{s.get("pnr_ball_handler", "None"):>8} '
            f'{s.get("passer", "None"):>8} '
            f'{s.get("off_dribble_shooter", "None"):>8} '
            f'{s.get("isolation_scorer", "None"):>8} '
            f'{raw["pnr_orchestration"]:>13.1f}'
        )

    print(f"\n--- Proposed coefficients ---")
    print(f"  shot_creation_pnr_orchestration: {PROPOSED_COEFFICIENTS['shot_creation_pnr_orchestration']}")
    print(f"  shot_creation_passer:            {PROPOSED_COEFFICIENTS.get('shot_creation_passer', 'N/A')}")
    print(f"  shot_creation_off_dribble:       {PROPOSED_COEFFICIENTS['shot_creation_off_dribble']}")
    print(f"  shot_creation_spacing:           {PROPOSED_COEFFICIENTS['shot_creation_spacing']}")
    print(f"  shot_creation_paint_touch:        {PROPOSED_COEFFICIENTS['shot_creation_paint_touch']}")
    print(f"  isolation_scorer:                 1.0 (unweighted)")

    # Export reference players to xlsx
    export_xlsx(ref_records, PROPOSED_COEFFICIENTS)


if __name__ == "__main__":
    main()
